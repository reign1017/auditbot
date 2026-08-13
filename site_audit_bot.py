#!/usr/bin/env python3
"""
SiteAuditBot - Comprehensive Sales Audit Report Generator

A CLI tool that takes a URL as input and generates a comprehensive
"Sales Audit Report" for a potential client.

Tasks:
1. Performance Check (Google PageSpeed Insights)
2. Technical Check (Schema & Meta)
3. AI Conversion Audit (XAI/Grok API or Gemini API)
"""

import argparse
import csv
import json
import os
import re
import socket
import subprocess
import sys
import time
import threading
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse, urljoin

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv

# Try to import extruct for enhanced schema detection
try:
    import extruct
    from w3lib.html import get_base_url
    EXTRACT_RICH_SCHEMA = True
except ImportError:
    extruct = None
    get_base_url = None
    EXTRACT_RICH_SCHEMA = False

# Try to import Wappalyzer for tech stack detection
try:
    from Wappalyzer import Wappalyzer, WebPage
    WAPPALYZER_AVAILABLE = True
except ImportError:
    Wappalyzer = None
    WebPage = None
    WAPPALYZER_AVAILABLE = False

# Try to import Playwright for axe-core accessibility
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    sync_playwright = None
    PLAYWRIGHT_AVAILABLE = False

# Try to import openpyxl for Excel support
try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to use system proxy settings if available
import urllib.request
try:
    # Get system proxy settings
    proxies = urllib.request.getproxies()
    if proxies:
        os.environ.update({
            'HTTP_PROXY': proxies.get('http', ''),
            'HTTPS_PROXY': proxies.get('https', ''),
        })
except:
    pass

# DNS preflight check - check DNS before attempting HTTP fetch
def _get_hostname(url: str) -> str:
    """Extract hostname from URL."""
    try:
        return urlparse(url).hostname or ""
    except Exception:
        return ""

def _preflight_dns(url: str) -> bool:
    """Check DNS resolution before attempting HTTP fetch."""
    host = _get_hostname(url)
    if not host:
        return False
    try:
        socket.getaddrinfo(host, 443)
        return True
    except socket.gaierror:
        return False

# Cache helper functions
def _load_audit_cache(path: str) -> dict:
    """Load audit cache from JSON file."""
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                obj = json.load(f)
                return obj if isinstance(obj, dict) else {}
    except Exception:
        pass
    return {}

def _save_audit_cache(path: str, cache: dict) -> None:
    """Save audit cache to JSON file."""
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _cache_get(cache: dict, key: str, ttl_seconds: float) -> dict:
    """Get cached value if not expired."""
    v = cache.get(key)
    if not isinstance(v, dict):
        return None
    if v.get("v") != AUDIT_CACHE_VERSION:
        return None
    ts = v.get("ts")
    if not isinstance(ts, (int, float)):
        return None
    if (time.time() - float(ts)) > float(ttl_seconds):
        return None
    return v

# Email extraction
EMAIL_REGEX = re.compile(r'(?i)[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}')
BLOCKED_EMAIL_DOMAINS = {
    "example.com", "example.org", "example.net", "email.com",
    "youremail.com", "yourdomain.com", "domain.com", "company.com",
    "yourcompany.com", "test.com"
}

def extract_email(html: str) -> str:
    """Extract first valid email address from HTML."""
    matches = EMAIL_REGEX.findall(html)
    for m in matches:
        clean = m.strip(" \t\r\n\"'<>),.;:]}")
        l = clean.lower()
        if l.count("@") != 1:
            continue
        parts = l.split("@", 1)
        if len(parts) != 2:
            continue
        local, domain = parts
        if not local or not domain or "." not in domain:
            continue
        if local.startswith(("no-reply", "noreply", "do-not-reply", "donotreply")):
            continue
        if domain in BLOCKED_EMAIL_DOMAINS:
            continue
        if "example" in domain or "youremail" in domain or "yourdomain" in domain:
            continue
        return clean
    return ""

# Google Maps link extraction
def extract_google_maps_link(soup: BeautifulSoup, base_url: str, schema_objs: list) -> str:
    """Extract Google Maps link from schema or HTML."""
    map_hosts = ("google.com/maps", "maps.google.com", "g.page", "goo.gl/maps", "maps.app.goo.gl")
    
    # Check schema first
    def walk_schema(obj):
        if isinstance(obj, dict):
            for k in ("hasMap", "map", "url", "sameAs", "@id"):
                v = obj.get(k)
                if isinstance(v, str):
                    if any(h in v.lower() for h in map_hosts):
                        return v
                elif isinstance(v, list):
                    for it in v:
                        if isinstance(it, str) and any(h in it.lower() for h in map_hosts):
                            return it
            for v in obj.values():
                if isinstance(v, (dict, list)):
                    result = walk_schema(v)
                    if result:
                        return result
        elif isinstance(obj, list):
            for it in obj:
                result = walk_schema(it)
                if result:
                    return result
        return None
    
    for obj in schema_objs:
        result = walk_schema(obj)
        if result:
            return result
    
    # Fallback to HTML
    for a in soup.select("a[href]"):
        href = (a.get("href") or "").strip()
        if not href:
            continue
        href_l = href.lower()
        if any(h in href_l for h in map_hosts):
            if href.startswith("/"):
                return urljoin(base_url, href)
            return href
    return ""

# City/Region extraction from schema
def extract_location_from_schema(schema_objs: list) -> tuple:
    """Extract city and region from schema. Returns (city, region)."""
    cities = set()
    regions = set()
    
    def walk(obj):
        if isinstance(obj, dict):
            addr = obj.get("address")
            if isinstance(addr, dict):
                loc = addr.get("addressLocality")
                reg = addr.get("addressRegion")
                if isinstance(loc, str) and loc.strip():
                    cities.add(loc.strip())
                if isinstance(reg, str) and reg.strip():
                    regions.add(reg.strip())
            elif isinstance(addr, list):
                for a in addr:
                    if isinstance(a, dict):
                        loc = a.get("addressLocality")
                        reg = a.get("addressRegion")
                        if isinstance(loc, str) and loc.strip():
                            cities.add(loc.strip())
                        if isinstance(reg, str) and reg.strip():
                            regions.add(reg.strip())
            for v in obj.values():
                if isinstance(v, (dict, list)):
                    walk(v)
        elif isinstance(obj, list):
            for it in obj:
                walk(it)
    
    for obj in schema_objs:
        walk(obj)
    
    city = "; ".join(sorted(cities)) if cities else ""
    region = "; ".join(sorted(regions)) if regions else ""
    return city, region

# Better name extraction
def extract_business_name(soup: BeautifulSoup, schema_objs: list, url: str) -> str:
    """Extract business name from schema, title, or domain."""
    # Try schema first
    def walk_schema(obj):
        names = []
        if isinstance(obj, dict):
            for k in ("name", "legalName"):
                v = obj.get(k)
                if isinstance(v, str) and v.strip() and len(v.strip()) >= 3:
                    if v.strip().lower() not in {"home", "homepage", "website"}:
                        names.append(v.strip())
            for v in obj.values():
                if isinstance(v, (dict, list)):
                    names.extend(walk_schema(v))
        elif isinstance(obj, list):
            for it in obj:
                names.extend(walk_schema(it))
        return names
    
    for obj in schema_objs:
        names = walk_schema(obj)
        if names:
            return names[0]
    
    # Fallback to HTML title
    if soup.title and soup.title.string:
        title = soup.title.string.strip()
        # Clean common separators
        cleaned = re.split(r'\s*[|\-–—]\s*', title)[0].strip()
        if cleaned:
            return cleaned
    
    # Fallback to domain
    host = _get_hostname(url)
    if host:
        base = host.split(".")[0]
        return base.replace("-", " ").strip().title() or host
    
    return "Unknown"


# --- Free digital presence checks (DIGITAL_PRESENCE_STACK) ---

def check_security_headers(response) -> dict:
    """Check critical security headers. Uses requests response (headers lowercased)."""
    h = getattr(response, "headers", {}) or {}
    critical = [
        ("strict-transport-security", "HSTS"),
        ("x-content-type-options", "X-Content-Type-Options"),
        ("x-frame-options", "X-Frame-Options"),
    ]
    optional = [
        ("content-security-policy", "CSP"),
        ("referrer-policy", "Referrer-Policy"),
    ]
    present = []
    missing = []
    for key, name in critical + optional:
        v = h.get(key)
        if v:
            present.append(name)
        else:
            missing.append(name)
    pass_ = all(
        name in present
        for _, name in critical
    )
    return {"pass": pass_, "present": present, "missing": missing}


GTM_RE = re.compile(r"GTM-[A-Z0-9]+", re.I)
GA4_RE = re.compile(r"G-[A-Z0-9]+", re.I)

def detect_gtm_ga4_clarity(html: str) -> dict:
    """Detect GTM, GA4, Microsoft Clarity from HTML."""
    html = (html or "").lower()
    out = {
        "gtm_present": False,
        "gtm_id": "",
        "ga4_present": False,
        "ga4_id": "",
        "clarity_present": False,
    }
    if "googletagmanager.com/gtm.js" in html or "googletagmanager.com/gtag/js" in html:
        m = GTM_RE.search(html)
        if m:
            out["gtm_present"] = True
            out["gtm_id"] = m.group(0)
    if "googletagmanager.com/gtag/js" in html or "gtag(" in html:
        m = GA4_RE.search(html)
        if m:
            out["ga4_present"] = True
            out["ga4_id"] = m.group(0)
    if "clarity.ms" in html or "clarity.microsoft.com" in html:
        out["clarity_present"] = True
    return out


def detect_wappalyzer(url: str) -> dict:
    """Detect tech stack via Wappalyzer (CMS, analytics, framework, etc.)."""
    out = {
        "cms": "",
        "analytics_tool": "",
        "chat_widget": "",
        "framework": "",
        "tech_stack_summary": "",
    }
    if not WAPPALYZER_AVAILABLE or not Wappalyzer or not WebPage:
        return out
    try:
        w = Wappalyzer.latest()
        page = WebPage.new_from_url(url, headers={"User-Agent": USER_AGENT})
        apps = w.analyze(page)
        if not apps:
            return out
        apps_list = list(apps) if isinstance(apps, (set, list)) else [str(a) for a in apps]
        out["tech_stack_summary"] = ", ".join(apps_list[:15])  # cap for Excel
        # Heuristic mapping: common CMS / analytics / chat
        lower = " ".join(apps_list).lower()
        cms_candidates = ["wordpress", "drupal", "joomla", "wix", "squarespace", "webflow", "shopify", "ghost"]
        for c in cms_candidates:
            if c in lower:
                out["cms"] = c
                break
        analytics_candidates = ["google analytics", "ga4", "gtag", "matomo", "plausible", "fathom", "hotjar", "clarity"]
        for a in analytics_candidates:
            if a in lower:
                out["analytics_tool"] = a
                break
        if "intercom" in lower or "drift" in lower or "crisp" in lower or "tidio" in lower or "livechat" in lower or "hubspot" in lower:
            out["chat_widget"] = "Yes"
        if "react" in lower or "vue" in lower or "angular" in lower or "jquery" in lower:
            out["framework"] = "React" if "react" in lower else ("Vue" if "vue" in lower else ("Angular" if "angular" in lower else "jQuery"))
    except Exception:
        pass
    return out


def schema_validate_basic(schema_objs: list) -> dict:
    """Basic local schema validation: has @type or @graph, no obvious issues."""
    valid = True
    warnings = []
    for obj in schema_objs or []:
        if not isinstance(obj, dict):
            valid = False
            warnings.append("Non-object in schema")
            continue
        if "@type" not in obj and "@graph" not in obj:
            valid = False
            warnings.append("Block missing @type and @graph")
    return {"valid": valid, "warnings": warnings}


def fetch_crux(origin: str, api_key: str) -> dict:
    """Chrome UX Report API (free). Requires Chrome UX Report API enabled. Returns LCP/FID/CLS p75."""
    out = {"lcp_p75": None, "fid_p75": None, "cls_p75": None, "error": None}
    url = "https://chromeuxreport.googleapis.com/v1/records:queryRecord"
    try:
        r = requests.post(
            url,
            params={"key": api_key},
            json={"origin": origin},
            timeout=15,
        )
        r.raise_for_status()
        data = r.json()
        rec = data.get("record") or {}
        metrics = rec.get("metrics") or {}
        for name, key in [("lcp_p75", "largest_contentful_paint"), ("fid_p75", "first_input_delay"), ("cls_p75", "cumulative_layout_shift")]:
            m = metrics.get(key) or {}
            p = m.get("percentiles") or {}
            val = p.get("p75")
            if val is not None:
                out[name] = val
        return out
    except Exception as e:
        out["error"] = str(e)
        return out


def fetch_ssl_labs_grade(host: str, timeout_s: int = 90, poll_interval: int = 5) -> dict:
    """SSL Labs API (free, async). Poll until READY or timeout. Returns grade."""
    out = {"grade": None, "error": None}
    base = "https://api.ssllabs.com/api/v3"
    try:
        r = requests.get(f"{base}/analyze", params={"host": host}, timeout=30)
        r.raise_for_status()
        data = r.json()
        if data.get("status") == "READY":
            eps = data.get("endpoints") or []
            for ep in eps:
                if isinstance(ep, dict) and ep.get("grade"):
                    out["grade"] = ep.get("grade")
                    break
            return out
        elapsed = 0
        while elapsed < timeout_s:
            time.sleep(min(poll_interval, timeout_s - elapsed))
            elapsed += poll_interval
            r = requests.get(f"{base}/analyze", params={"host": host}, timeout=30)
            r.raise_for_status()
            data = r.json()
            st = data.get("status")
            if st == "READY":
                eps = data.get("endpoints") or []
                for ep in eps:
                    if isinstance(ep, dict) and ep.get("grade"):
                        out["grade"] = ep.get("grade")
                        break
                return out
            if st == "ERROR":
                out["error"] = data.get("statusMessage") or "SSL Labs error"
                return out
        out["error"] = "timeout"
        return out
    except Exception as e:
        out["error"] = str(e)
        return out


# Optional: Only import if using Gemini
try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

# Load environment variables from .env file if it exists
load_dotenv()

# Configure console output for Windows Unicode support
def _configure_console_output():
    """Prevent UnicodeEncodeError on Windows terminals (e.g., emojis)."""
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

_configure_console_output()

# Standard User-Agent to avoid 403 blocks
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# Cache system for performance
AUDIT_CACHE_VERSION = 1
audit_cache = {}
cache_lock = threading.Lock()


def get_api_keys():
    """Get API keys from environment or prompt user for input."""
    # Try to load from environment first
    pagespeed_key = os.getenv("PAGESPEED_API_KEY")
    mistral_key = os.getenv("MISTRAL_API_KEY")
    xai_key = os.getenv("XAI_API_KEY")
    gemini_key = os.getenv("GEMINI_API_KEY")
    
    # Prompt for missing keys
    if not pagespeed_key:
        pagespeed_key = input("Enter your Google PageSpeed Insights API Key: ").strip()
    
    # Prefer Mistral, then XAI, then Gemini
    if not mistral_key and not xai_key and not gemini_key:
        choice = input("Which AI API do you want to use? (1) Mistral AI (2) XAI/Grok (3) Gemini [1]: ").strip() or "1"
        if choice == "1":
            mistral_key = input("Enter your Mistral AI API Key: ").strip()
        elif choice == "2":
            xai_key = input("Enter your XAI API Key: ").strip()
        else:
            gemini_key = input("Enter your Google Gemini API Key: ").strip()
    
    return pagespeed_key, mistral_key, xai_key, gemini_key


def task1_performance_check(url: str, api_key: str) -> dict:
    """
    Task 1: The Performance Check (Google PageSpeed)
    
    Uses the Google PageSpeed Insights API to get mobile performance score.
    Flags as CRITICAL FAIL if score is under 50.
    """
    print("\n[Task 1] Running Performance Check...")
    
    try:
        api_url = "https://www.googleapis.com/pagespeedonline/v5/runPagespeed"
        params = {
            "url": url,
            "strategy": "mobile",
            "key": api_key
        }
        
        # PageSpeed API can be slow, so we retry with increasing timeouts
        max_retries = 2
        timeout = 60
        
        for attempt in range(max_retries + 1):
            try:
                response = requests.get(api_url, params=params, timeout=timeout)
                response.raise_for_status()
                break  # Success, exit retry loop
            except requests.exceptions.Timeout:
                if attempt < max_retries:
                    timeout += 30  # Increase timeout for retry
                    print(f"  ⚠ Timeout, retrying with {timeout}s timeout... (attempt {attempt + 2}/{max_retries + 1})")
                    continue
                else:
                    raise requests.exceptions.Timeout(f"PageSpeed API timed out after {max_retries + 1} attempts. The API may be slow or unavailable.")
        
        data = response.json()
        
        # Extract performance score
        lighthouse_result = data.get("lighthouseResult", {})
        categories = lighthouse_result.get("categories", {})
        performance = categories.get("performance", {})
        score = performance.get("score", 0)
        
        # Convert to 0-100 scale (API returns 0-1)
        performance_score = int(score * 100) if score else 0
        
        status = "CRITICAL FAIL" if performance_score < 50 else "PASS"
        
        # Lighthouse accessibility score (free - from same PageSpeed response)
        a11y_cat = categories.get("accessibility", {})
        a11y_score_raw = a11y_cat.get("score")
        accessibility_score = int(a11y_score_raw * 100) if a11y_score_raw is not None else None
        
        result = {
            "performance_score": performance_score,
            "status": status,
            "raw_score": score,
            "accessibility_score": accessibility_score,
            "crux": None
        }
        
        # Chrome UX Report (free) - real-user Core Web Vitals
        try:
            parsed = urlparse(url)
            origin = f"{parsed.scheme or 'https'}://{parsed.netloc or _get_hostname(url)}"
            crux = fetch_crux(origin, api_key)
            if crux.get("error"):
                pass  # API often not enabled; skip silently
            else:
                result["crux"] = crux
                parts = []
                if crux.get("lcp_p75") is not None:
                    parts.append(f"LCP p75={crux['lcp_p75']}ms")
                if crux.get("fid_p75") is not None:
                    parts.append(f"FID p75={crux['fid_p75']}ms")
                if crux.get("cls_p75") is not None:
                    parts.append(f"CLS p75={crux['cls_p75']}")
                if parts:
                    print(f"  ✓ CrUX (real users): {', '.join(parts)}")
        except Exception:
            pass
        
        print(f"  ✓ Performance Score (Mobile): {performance_score}/100")
        if accessibility_score is not None:
            print(f"  ✓ Accessibility Score (Lighthouse): {accessibility_score}/100")
        if status == "CRITICAL FAIL":
            print(f"  ⚠ WARNING: Performance score is below 50!")
        
        return result
        
    except requests.exceptions.RequestException as e:
        print(f"  ✗ Error fetching PageSpeed data: {e}")
        return {
            "performance_score": None,
            "status": "ERROR",
            "accessibility_score": None,
            "error": str(e)
        }
    except (KeyError, ValueError) as e:
        print(f"  ✗ Error parsing PageSpeed response: {e}")
        return {
            "performance_score": None,
            "status": "ERROR",
            "accessibility_score": None,
            "error": str(e)
        }


def _extract_schema_objects_from_html(html: str, base_url: str) -> list:
    """Extract schema objects from HTML using extruct (JSON-LD + Microdata + RDFa) or fallback to JSON-LD only."""
    html = html or ""
    base_url = base_url or ""
    
    if EXTRACT_RICH_SCHEMA and extruct and get_base_url:
        try:
            extracted = extruct.extract(
                html,
                base_url=get_base_url(html, base_url) or base_url,
                syntaxes=["json-ld", "microdata", "rdfa"],
                uniform=True,
            )
            out = []
            for k in ("json-ld", "microdata", "rdfa"):
                v = extracted.get(k) if isinstance(extracted, dict) else None
                if isinstance(v, list) and v:
                    out.extend(v)
            return out
        except Exception:
            pass  # Fall back to JSON-LD parsing below
    
    # Fallback to JSON-LD only
    soup = BeautifulSoup(html, "html.parser")
    objs = []
    scripts = soup.find_all("script", attrs={"type": re.compile(r"application/ld\+json", re.I)})
    for s in scripts:
        raw = s.string or s.get_text() or ""
        if not raw.strip():
            continue
        try:
            # Try direct parse
            objs.append(json.loads(raw))
        except json.JSONDecodeError:
            # Try extracting largest JSON object/array
            first_brace = min([i for i in [raw.find("{"), raw.find("[")] if i != -1], default=-1)
            if first_brace != -1:
                trimmed = raw[first_brace:]
                last_close = max(trimmed.rfind("}"), trimmed.rfind("]"))
                if last_close != -1:
                    try:
                        objs.append(json.loads(trimmed[:last_close + 1]))
                    except:
                        pass
    return objs

def _iter_jsonld_entities(obj) -> list:
    """Recursively extract all JSON-LD entities (objects with @type)."""
    entities = []
    if isinstance(obj, list):
        for item in obj:
            entities.extend(_iter_jsonld_entities(item))
    elif isinstance(obj, dict):
        # @graph commonly contains the real entities
        graph = obj.get("@graph")
        if isinstance(graph, list):
            for item in graph:
                entities.extend(_iter_jsonld_entities(item))
        if "@type" in obj:
            entities.append(obj)
        # Also traverse nested dict/list values
        for v in obj.values():
            if isinstance(v, (dict, list)):
                entities.extend(_iter_jsonld_entities(v))
    return entities

def _normalize_types(t) -> list:
    """Normalize @type to list of strings."""
    if isinstance(t, str):
        return [t.strip()]
    if isinstance(t, list):
        return [item.strip() for item in t if isinstance(item, str)]
    return []

def _is_nonempty(value) -> bool:
    """Check if value is non-empty."""
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    if isinstance(value, (list, tuple, set, dict)):
        return len(value) > 0
    return True

def task2_technical_check(url: str) -> dict:
    """
    Task 2: Enhanced Technical Check (Schema & Meta)
    
    Enhanced with:
    - Rich schema detection (JSON-LD + Microdata + RDFa via extruct)
    - Comprehensive AI visibility scoring (0-5)
    - Email extraction
    - Google Maps link extraction
    - City/Region extraction
    - Better name extraction
    """
    print("\n[Task 2] Running Enhanced Technical Check...")
    
    result = {
        "jsonld_exists": False,
        "schema_types_found": [],
        "title": None,
        "business_name": None,
        "email": "",
        "google_maps_link": "",
        "city": "",
        "region": "",
        "ai_visibility_score": 0,
        "ai_visibility_max": 5,
        "missing_opportunities": [],
        "security_headers": None,
        "gtm_present": False,
        "gtm_id": "",
        "ga4_present": False,
        "ga4_id": "",
        "clarity_present": False,
        "schema_valid": True,
        "schema_warnings": [],
        "cms": "",
        "analytics_tool": "",
        "chat_widget": "",
        "framework": "",
        "tech_stack_summary": "",
        "error": None
    }
    
    try:
        # DNS preflight check
        if not _preflight_dns(url):
            host = _get_hostname(url) or "(unknown host)"
            error_msg = f"Could not resolve domain (DNS): {host}"
            print(f"  ✗ {error_msg}")
            print("   - Double-check the spelling of the domain")
            print("   - Try adding/removing 'www.' (e.g., https://www.example.com)")
            result["error"] = error_msg
            return result
        
        # Fetch with HTTP/HTTPS fallback
        headers = {"User-Agent": USER_AGENT}
        try:
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
        except requests.exceptions.RequestException:
            # Try HTTP if HTTPS failed
            if url.startswith("https://"):
                http_url = url.replace("https://", "http://")
                try:
                    response = requests.get(http_url, headers=headers, timeout=30)
                    response.raise_for_status()
                    url = http_url  # Update URL for consistency
                except:
                    raise
        
        soup = BeautifulSoup(response.content, "html.parser")
        final_url = response.url or url
        
        # Security headers (free)
        sh = check_security_headers(response)
        result["security_headers"] = sh
        if sh["pass"]:
            print(f"  ✓ Security headers: PASS")
        else:
            print(f"  ⚠ Security headers: FAIL (missing: {', '.join(sh['missing'][:3])})")
        
        # GTM / GA4 / Clarity detection (free)
        tech = detect_gtm_ga4_clarity(response.text)
        result["gtm_present"] = tech["gtm_present"]
        result["gtm_id"] = tech["gtm_id"]
        result["ga4_present"] = tech["ga4_present"]
        result["ga4_id"] = tech["ga4_id"]
        result["clarity_present"] = tech["clarity_present"]
        if tech["gtm_id"]:
            print(f"  ✓ GTM: {tech['gtm_id']}")
        if tech["ga4_id"]:
            print(f"  ✓ GA4: {tech['ga4_id']}")
        if tech["clarity_present"]:
            print(f"  ✓ Clarity: present")
        
        # Wappalyzer tech stack (free)
        wap = detect_wappalyzer(final_url)
        result["cms"] = wap.get("cms", "")
        result["analytics_tool"] = wap.get("analytics_tool", "")
        result["chat_widget"] = wap.get("chat_widget", "")
        result["framework"] = wap.get("framework", "")
        result["tech_stack_summary"] = wap.get("tech_stack_summary", "")
        if wap.get("tech_stack_summary"):
            print(f"  ✓ Tech stack: {wap['tech_stack_summary'][:60]}...")
        if wap.get("cms"):
            print(f"  ✓ CMS: {wap['cms']}")
        
        # Extract schema using enhanced method
        schema_objs = _extract_schema_objects_from_html(response.text, final_url)
        result["jsonld_exists"] = len(schema_objs) > 0
        
        # Basic schema validation (free)
        sv = schema_validate_basic(schema_objs)
        result["schema_valid"] = sv["valid"]
        result["schema_warnings"] = sv["warnings"]
        if sv["warnings"]:
            print(f"  ⚠ Schema warnings: {'; '.join(sv['warnings'][:3])}")
        
        if result["jsonld_exists"]:
            print(f"  ✓ Schema found: Yes ({len(schema_objs)} block(s))")
        else:
            print("  ✗ Schema found: No")
        
        # Extract all entities
        entities = []
        for obj in schema_objs:
            entities.extend(_iter_jsonld_entities(obj))
        
        # Check for target types
        target_types = ["Attorney", "LegalService", "LocalBusiness", "HomeAndConstructionBusiness", 
                       "LawFirm", "Organization"]
        found_types = []
        relevant_entities = []
        
        for entity in entities:
            types = _normalize_types(entity.get("@type"))
            types_lower = [t.lower() for t in types]
            for t in types:
                if t in target_types or any(tt.lower() in types_lower for tt in target_types):
                    found_types.append(t)
                    if entity not in relevant_entities:
                        relevant_entities.append(entity)
        
        result["schema_types_found"] = list(set(found_types))
        
        if result["schema_types_found"]:
            print(f"  ✓ Target schema types found: {', '.join(result['schema_types_found'])}")
        else:
            print("  ✗ Target schema types found: None")
        
        # Comprehensive AI Visibility Scoring (0-5)
        score = 0
        max_score = 5
        missing = []
        
        def any_entity_has(prop: str) -> bool:
            for ent in relevant_entities:
                if _is_nonempty(ent.get(prop)):
                    return True
            return False
        
        # Check 1: Correct @type (LawFirm/Attorney/LegalService)
        has_lawfirm_type = any(
            any(t.lower() in ["lawfirm", "attorney", "legalservice"] 
                for t in _normalize_types(ent.get("@type")))
            for ent in relevant_entities
        )
        if has_lawfirm_type:
            score += 1
        else:
            missing.append("Add @type 'LawFirm', 'Attorney', or 'LegalService'")
        
        # Check 2: knowsAbout (Critical for AI understanding services)
        if any_entity_has("knowsAbout"):
            score += 1
        else:
            missing.append("Add 'knowsAbout' with specific services/topics")
        
        # Check 3: areaServed (Critical for local SEO)
        if any_entity_has("areaServed"):
            score += 1
        else:
            missing.append("Add 'areaServed' (cities/regions served)")
        
        # Check 4: sameAs (Social proof)
        if any_entity_has("sameAs"):
            score += 1
        else:
            missing.append("Add 'sameAs' with authoritative profile links")
        
        # Check 5: aggregateRating (Review stars)
        if any_entity_has("aggregateRating"):
            score += 1
        else:
            missing.append("Add 'aggregateRating' (with ratingValue/reviewCount)")
        
        result["ai_visibility_score"] = score
        result["ai_visibility_max"] = max_score
        result["missing_opportunities"] = missing
        
        if score >= 4:
            status = "EXCELLENT"
        elif score >= 2:
            status = "BASIC"
        else:
            status = "POOR"
        print(f"  ✓ AI Visibility Score: {score}/{max_score} ({status})")
        
        # Extract business name
        result["business_name"] = extract_business_name(soup, schema_objs, final_url)
        if result["business_name"]:
            print(f"  ✓ Business Name: {result['business_name']}")
        
        # Extract email
        result["email"] = extract_email(response.text)
        if not result["email"]:
            # Try contact page
            try:
                contact_url = urljoin(final_url, "/contact")
                contact_resp = requests.get(contact_url, headers=headers, timeout=15)
                if contact_resp.status_code == 200:
                    result["email"] = extract_email(contact_resp.text)
            except:
                pass
        if result["email"]:
            print(f"  ✓ Email: {result['email']}")
        
        # Extract Google Maps link
        result["google_maps_link"] = extract_google_maps_link(soup, final_url, schema_objs)
        if result["google_maps_link"]:
            print(f"  ✓ Google Maps: Found")
        
        # Extract location
        city, region = extract_location_from_schema(schema_objs)
        result["city"] = city
        result["region"] = region
        if city or region:
            print(f"  ✓ Location: {city or 'N/A'}, {region or 'N/A'}")
        
        # Capture title
        title_tag = soup.find("title")
        if title_tag:
            result["title"] = title_tag.get_text(strip=True)
        else:
            result["title"] = "Not found"
        
        return result
        
    except requests.exceptions.RequestException as e:
        error_msg = f"Error fetching URL: {e}"
        print(f"  ✗ {error_msg}")
        result["error"] = error_msg
        return result
    except Exception as e:
        error_msg = f"Error parsing HTML: {e}"
        print(f"  ✗ {error_msg}")
        result["error"] = error_msg
        return result


def extract_visible_text(html_content: str) -> str:
    """Extract visible text from HTML, cleaning out HTML tags."""
    soup = BeautifulSoup(html_content, "html.parser")
    
    # Remove script and style elements
    for script in soup(["script", "style"]):
        script.decompose()
    
    # Get text and clean it up
    text = soup.get_text(separator=" ", strip=True)
    
    # Clean up whitespace
    text = re.sub(r"\s+", " ", text)
    
    return text.strip()


def task3_ai_conversion_audit_xai(url: str, api_key: str) -> dict:
    """
    Task 3: The AI Conversion Audit (XAI/Grok API)
    
    Extracts all visible text from the homepage, sends first 2,000 characters
    to XAI Grok API, and gets CRO feedback.
    """
    print("\n[Task 3] Running AI Conversion Audit (XAI/Grok)...")
    
    result = {
        "conversion_feedback": None,
        "error": None
    }
    
    try:
        # Fetch the page
        headers = {"User-Agent": USER_AGENT}
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        
        # Extract visible text
        visible_text = extract_visible_text(response.text)
        
        # Get first 1,500 characters (reduced for cost efficiency)
        # This is ~375 tokens input, keeping cost under $0.01 per audit
        text_sample = visible_text[:1500]
        
        if not text_sample:
            result["error"] = "No visible text found on page"
            print("  ✗ No visible text found on page")
            return result
        
        print(f"  ✓ Extracted {len(text_sample)} characters of visible text")
        print(f"  💰 Estimated cost: ~$0.01 (well under your $0.20 limit)")
        
        # Create the prompt
        prompt = (
            "Act as a harsh CRO expert. Analyze this homepage text. "
            "Give me 3 bullet points on why a customer would leave this site without buying. "
            "Be specific and brutal."
        )
        
        full_prompt = f"{prompt}\n\nHomepage text:\n{text_sample}"
        
        # Call XAI API - OpenAI-compatible endpoint
        # XAI API is compatible with OpenAI format
        api_url = "https://api.x.ai/v1/chat/completions"
        headers_api = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        
        # Try different XAI models in order (newest first)
        models_to_try = [
            "grok-2-1212",      # Latest model
            "grok-2-vision-1212",
            "grok-beta",
            "grok-2-beta",
            "grok-3-beta",
            "grok-2",
        ]
        
        last_error = None
        for model_name in models_to_try:
            try:
                payload = {
                    "model": model_name,
                    "messages": [
                        {
                            "role": "user",
                            "content": full_prompt
                        }
                    ],
                    "temperature": 0.7,
                    "max_tokens": 300  # Reduced from 500 to keep costs low
                }
                
                response_xai = requests.post(api_url, headers=headers_api, json=payload, timeout=60)
                
                if response_xai.status_code == 200:
                    # Success!
                    data = response_xai.json()
                    result["conversion_feedback"] = data["choices"][0]["message"]["content"].strip()
                    print(f"  ✓ Received CRO feedback from XAI/Grok (using model: {model_name})")
                    return result
                elif response_xai.status_code == 404:
                    # Model not found, try next one
                    last_error = f"Model '{model_name}' not found (404)"
                    continue
                else:
                    # Other error, raise it
                    response_xai.raise_for_status()
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 404:
                    # Model not found, try next
                    last_error = f"Model '{model_name}' not found (404)"
                    continue
                else:
                    # Other HTTP error, raise it
                    raise
        
        # If we get here, all models failed
        raise Exception(f"All XAI models failed. Last error: {last_error}. Please check https://console.x.ai/ for available models and API access status.")
        
    except requests.exceptions.RequestException as e:
        error_msg = f"Error with XAI API: {e}"
        print(f"  ✗ {error_msg}")
        result["error"] = error_msg
        return result
    except (KeyError, ValueError) as e:
        error_msg = f"Error parsing XAI API response: {e}"
        print(f"  ✗ {error_msg}")
        result["error"] = error_msg
        return result
    except Exception as e:
        error_msg = f"Error with XAI API: {e}"
        print(f"  ✗ {error_msg}")
        result["error"] = error_msg
        return result


def task3_ai_conversion_audit_mistral(url: str, api_key: str) -> dict:
    """
    Task 3: The AI Conversion Audit (Mistral AI)
    
    Extracts all visible text from the homepage, sends first 1,500 characters
    to Mistral AI API, and gets CRO feedback.
    """
    print("\n[Task 3] Running AI Conversion Audit (Mistral AI)...")
    
    result = {
        "conversion_feedback": None,
        "error": None
    }
    
    try:
        # Fetch the page
        headers = {"User-Agent": USER_AGENT}
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        
        # Extract visible text
        visible_text = extract_visible_text(response.text)
        
        # Get first 1,500 characters (reduced for cost efficiency)
        text_sample = visible_text[:1500]
        
        if not text_sample:
            result["error"] = "No visible text found on page"
            print("  ✗ No visible text found on page")
            return result
        
        print(f"  ✓ Extracted {len(text_sample)} characters of visible text")
        print(f"  💰 Estimated cost: ~$0.01 (well under your $0.20 limit)")
        
        # Create the prompt
        prompt = (
            "Act as a harsh CRO expert. Analyze this homepage text. "
            "Give me 3 bullet points on why a customer would leave this site without buying. "
            "Be specific and brutal."
        )
        
        full_prompt = f"{prompt}\n\nHomepage text:\n{text_sample}"
        
        # Call Mistral AI API
        api_url = "https://api.mistral.ai/v1/chat/completions"
        headers_api = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        
        # Try models in order of preference (cheapest first)
        models_to_try = ["mistral-small", "mistral-tiny", "mistral-medium"]
        
        for model_name in models_to_try:
            try:
                payload = {
                    "model": model_name,
                    "messages": [
                        {
                            "role": "user",
                            "content": full_prompt
                        }
                    ],
                    "temperature": 0.7,
                    "max_tokens": 300  # Reduced to keep costs low
                }
                
                response_mistral = requests.post(api_url, headers=headers_api, json=payload, timeout=60)
                
                if response_mistral.status_code == 200:
                    data = response_mistral.json()
                    result["conversion_feedback"] = data["choices"][0]["message"]["content"].strip()
                    print(f"  ✓ Received CRO feedback from Mistral AI (using model: {model_name})")
                    return result
                elif response_mistral.status_code == 404:
                    # Model not found, try next one
                    continue
                else:
                    response_mistral.raise_for_status()
            except requests.exceptions.HTTPError:
                if model_name == models_to_try[-1]:
                    # Last model failed, raise the error
                    raise
                continue
        
        # If we get here, all models failed
        raise Exception("All Mistral models failed")
        
    except requests.exceptions.RequestException as e:
        error_msg = f"Error with Mistral AI API: {e}"
        print(f"  ✗ {error_msg}")
        result["error"] = error_msg
        return result
    except Exception as e:
        error_msg = f"Error with Mistral AI API: {e}"
        print(f"  ✗ {error_msg}")
        result["error"] = error_msg
        return result


def task3_ai_conversion_audit_gemini(url: str, api_key: str) -> dict:
    """
    Task 3: The AI Conversion Audit (Gemini API)
    
    Extracts all visible text from the homepage, sends first 2,000 characters
    to Google Gemini API, and gets CRO feedback.
    """
    print("\n[Task 3] Running AI Conversion Audit (Gemini)...")
    
    result = {
        "conversion_feedback": None,
        "error": None
    }
    
    if not GEMINI_AVAILABLE:
        result["error"] = "google-generativeai package not installed. Install with: pip install google-generativeai"
        print("  ✗ Gemini package not available")
        return result
    
    try:
        # Fetch the page
        headers = {"User-Agent": USER_AGENT}
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        
        # Extract visible text
        visible_text = extract_visible_text(response.text)
        
        # Get first 2,000 characters
        text_sample = visible_text[:2000]
        
        if not text_sample:
            result["error"] = "No visible text found on page"
            print("  ✗ No visible text found on page")
            return result
        
        print(f"  ✓ Extracted {len(text_sample)} characters of visible text")
        
        # Configure Gemini API
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-pro")
        
        # Create the prompt
        prompt = (
            "Act as a harsh CRO expert. Analyze this homepage text. "
            "Give me 3 bullet points on why a customer would leave this site without buying. "
            "Be specific and brutal."
        )
        
        # Send to Gemini
        full_prompt = f"{prompt}\n\nHomepage text:\n{text_sample}"
        response_gemini = model.generate_content(full_prompt)
        
        result["conversion_feedback"] = response_gemini.text.strip()
        
        print("  ✓ Received CRO feedback from Gemini")
        return result
        
    except requests.exceptions.RequestException as e:
        error_msg = f"Error fetching URL: {e}"
        print(f"  ✗ {error_msg}")
        result["error"] = error_msg
        return result
    except Exception as e:
        error_msg = f"Error with Gemini API: {e}"
        print(f"  ✗ {error_msg}")
        result["error"] = error_msg
        return result


def task4_ssl_security(host: str, timeout_s: int = 60) -> dict:
    """Task 4: SSL Labs grade (free). Async; polls until READY or timeout."""
    print("\n[Task 4] Running SSL / Security Check (SSL Labs)...")
    r = fetch_ssl_labs_grade(host, timeout_s=timeout_s)
    grade = r.get("grade")
    err = r.get("error")
    if grade:
        print(f"  ✓ SSL Grade: {grade}")
    elif err:
        print(f"  ⚠ SSL Labs: {err}")
    return {"ssl_grade": grade, "ssl_error": err}


TONE_PROMPT = (
    "Score this law firm homepage 1-10 for client-focus vs self-promotion. "
    "1 = very self-promotional (awards, credentials only). 10 = highly client-focused (solutions, clear CTAs). "
    "Reply with: 'Score: N' then one short paragraph explaining why."
)


def _call_tone_xai(text_sample: str, api_key: str) -> tuple:
    """Call XAI for tone score. Returns (score int or None, notes str, error str)."""
    api_url = "https://api.x.ai/v1/chat/completions"
    models = ["grok-2-1212", "grok-2-vision-1212", "grok-beta", "grok-2"]
    for model in models:
        try:
            r = requests.post(
                api_url,
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={
                    "model": model,
                    "messages": [{"role": "user", "content": f"{TONE_PROMPT}\n\nHomepage text:\n{text_sample[:1500]}"}],
                    "temperature": 0.3,
                    "max_tokens": 200,
                },
                timeout=60,
            )
            if r.status_code != 200:
                continue
            content = r.json()["choices"][0]["message"]["content"].strip()
            m = re.search(r"\b([1-9]|10)\b", content)
            score = int(m.group(1)) if m else None
            return (score, content, None)
        except Exception as e:
            continue
    return (None, "", "XAI tone call failed")


def _call_tone_mistral(text_sample: str, api_key: str) -> tuple:
    """Call Mistral for tone score. Returns (score, notes, error)."""
    api_url = "https://api.mistral.ai/v1/chat/completions"
    for model in ["mistral-small", "mistral-tiny"]:
        try:
            r = requests.post(
                api_url,
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={
                    "model": model,
                    "messages": [{"role": "user", "content": f"{TONE_PROMPT}\n\nHomepage text:\n{text_sample[:1500]}"}],
                    "temperature": 0.3,
                    "max_tokens": 200,
                },
                timeout=60,
            )
            if r.status_code != 200:
                continue
            content = r.json()["choices"][0]["message"]["content"].strip()
            m = re.search(r"\b([1-9]|10)\b", content)
            score = int(m.group(1)) if m else None
            return (score, content, None)
        except Exception:
            continue
    return (None, "", "Mistral tone call failed")


def _call_tone_gemini(text_sample: str, api_key: str) -> tuple:
    """Call Gemini for tone score. Returns (score, notes, error)."""
    if not GEMINI_AVAILABLE:
        return (None, "", "Gemini not installed")
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-pro")
        r = model.generate_content(f"{TONE_PROMPT}\n\nHomepage text:\n{text_sample[:1500]}")
        content = r.text.strip()
        m = re.search(r"\b([1-9]|10)\b", content)
        score = int(m.group(1)) if m else None
        return (score, content, None)
    except Exception as e:
        return (None, "", str(e))


def task_tone_score(url: str, api_key: str, provider: str) -> dict:
    """AI Tone / Client-Focus Score (1-10). Uses same AI as Task 3."""
    out = {"tone_score": None, "tone_notes": "", "error": None}
    try:
        headers = {"User-Agent": USER_AGENT}
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        text = extract_visible_text(resp.text)[:1500]
        if not text:
            out["error"] = "No visible text"
            return out
        if provider == "xai":
            score, notes, err = _call_tone_xai(text, api_key)
        elif provider == "mistral":
            score, notes, err = _call_tone_mistral(text, api_key)
        else:
            score, notes, err = _call_tone_gemini(text, api_key)
        if err:
            out["error"] = err
        out["tone_score"] = score
        out["tone_notes"] = (notes or "")[:500]
        return out
    except Exception as e:
        out["error"] = str(e)
        return out


def task_axe_accessibility(url: str) -> dict:
    """axe-core accessibility scan via Playwright. Returns violations count + top 5 issues."""
    out = {
        "a11y_violations_count": 0,
        "a11y_critical_issues": "",
        "a11y_top_5": "",
        "error": None,
    }
    if not PLAYWRIGHT_AVAILABLE or not sync_playwright:
        out["error"] = "Playwright not installed. Run: pip install playwright && playwright install chromium"
        return out
    try:
        with sync_playwright() as p:
            with p.chromium.launch(headless=True) as browser:
                context = browser.new_context(
                    user_agent=USER_AGENT,
                    ignore_https_errors=True,
                )
                page = context.new_page()
                page.goto(url, wait_until="domcontentloaded", timeout=20000)
                page.add_script_tag(url="https://cdnjs.cloudflare.com/ajax/libs/axe-core/4.9.0/axe.min.js")
                raw = page.evaluate("""async () => {
                    const r = await axe.run();
                    return { violations: r.violations };
                }""")
        violations = raw.get("violations") or []
        out["a11y_violations_count"] = len(violations)
        top = []
        for v in violations[:5]:
            name = v.get("id") or v.get("help", "Unknown")
            impact = v.get("impact", "")
            nodes = len(v.get("nodes") or [])
            top.append(f"{name} (impact: {impact}, nodes: {nodes})")
        out["a11y_top_5"] = "; ".join(top)
        out["a11y_critical_issues"] = "; ".join(top[:3])
        return out
    except Exception as e:
        out["error"] = str(e)
        return out


def generate_report(url: str, task1_result: dict, task2_result: dict, task3_result: dict, task4_result: dict = None, tone_result: dict = None, axe_result: dict = None) -> str:
    """Generate a formatted report string."""
    domain = urlparse(url).netloc.replace("www.", "")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    task4_result = task4_result or {}
    
    report = f"""
{'='*70}
SITE AUDIT REPORT
{'='*70}
URL: {url}
Domain: {domain}
Generated: {timestamp}

{'='*70}
TASK 1: PERFORMANCE CHECK (Google PageSpeed Insights)
{'='*70}
"""
    
    if task1_result.get("error"):
        report += f"Status: ERROR\nError: {task1_result['error']}\n"
    else:
        score = task1_result.get("performance_score", "N/A")
        status = task1_result.get("status", "UNKNOWN")
        report += f"Performance Score (Mobile): {score}/100\n"
        report += f"Status: {status}\n"
        a11y = task1_result.get("accessibility_score")
        if a11y is not None:
            report += f"Accessibility Score (Lighthouse): {a11y}/100\n"
        crux = task1_result.get("crux") or {}
        if crux and not crux.get("error"):
            parts = []
            if crux.get("lcp_p75") is not None:
                parts.append(f"LCP p75={crux['lcp_p75']}ms")
            if crux.get("fid_p75") is not None:
                parts.append(f"FID p75={crux['fid_p75']}ms")
            if crux.get("cls_p75") is not None:
                parts.append(f"CLS p75={crux['cls_p75']}")
            if parts:
                report += f"CrUX (real users): {', '.join(parts)}\n"
        if status == "CRITICAL FAIL":
            report += "⚠ WARNING: Performance score is below 50. This is a critical issue.\n"
    
    report += f"""
{'='*70}
TASK 2: TECHNICAL CHECK (Schema & Meta)
{'='*70}
"""
    
    if task2_result.get("error"):
        report += f"Error: {task2_result['error']}\n"
    else:
        report += f"JSON-LD Schema Exists: {'Yes' if task2_result.get('jsonld_exists') else 'No'}\n"
        
        schema_types = task2_result.get("schema_types_found", [])
        if schema_types:
            report += f"Target Schema Types Found: {', '.join(schema_types)}\n"
        else:
            report += "Target Schema Types Found: None\n"
        
        # AI Visibility Score
        ai_score = task2_result.get("ai_visibility_score", 0)
        ai_max = task2_result.get("ai_visibility_max", 5)
        ai_status = "EXCELLENT" if ai_score >= 4 else ("BASIC" if ai_score >= 2 else "POOR")
        report += f"AI Visibility Score: {ai_score}/{ai_max} ({ai_status})\n"
        
        missing = task2_result.get("missing_opportunities", [])
        if missing:
            report += f"\nMissing Opportunities:\n"
            for item in missing:
                report += f"  - {item}\n"
        
        # Business Name
        business_name = task2_result.get("business_name")
        if business_name:
            report += f"\nBusiness Name: {business_name}\n"
        
        # Contact Info
        email = task2_result.get("email")
        if email:
            report += f"Email: {email}\n"
        
        google_maps = task2_result.get("google_maps_link")
        if google_maps:
            report += f"Google Maps: {google_maps}\n"
        
        # Location
        city = task2_result.get("city")
        region = task2_result.get("region")
        if city or region:
            report += f"Location: {city or 'N/A'}, {region or 'N/A'}\n"
        
        title = task2_result.get("title")
        if title:
            report += f"\nTitle Tag: {title}\n"
        else:
            report += "\nTitle Tag: Not found\n"
        
        # Free checks: security headers, GTM/GA4/Clarity, schema validation
        sh = task2_result.get("security_headers")
        if isinstance(sh, dict):
            report += f"\nSecurity Headers: {'PASS' if sh.get('pass') else 'FAIL'}\n"
            if sh.get("missing"):
                report += f"  Missing: {', '.join(sh['missing'])}\n"
        if task2_result.get("gtm_id"):
            report += f"GTM: {task2_result['gtm_id']}\n"
        if task2_result.get("ga4_id"):
            report += f"GA4: {task2_result['ga4_id']}\n"
        if task2_result.get("clarity_present"):
            report += "Clarity: present\n"
        report += f"Schema Valid: {'Yes' if task2_result.get('schema_valid') else 'No'}\n"
        w = task2_result.get("schema_warnings") or []
        if w:
            report += f"Schema Warnings: {'; '.join(w[:5])}\n"
        if task2_result.get("cms"):
            report += f"CMS: {task2_result['cms']}\n"
        if task2_result.get("framework"):
            report += f"Framework: {task2_result['framework']}\n"
        if task2_result.get("tech_stack_summary"):
            report += f"Tech Stack: {task2_result['tech_stack_summary'][:120]}...\n"
    
    report += f"""
{'='*70}
TASK 3: AI CONVERSION AUDIT
{'='*70}
"""
    
    if task3_result.get("error"):
        report += f"Error: {task3_result['error']}\n"
    else:
        feedback = task3_result.get("conversion_feedback")
        if feedback:
            report += f"{feedback}\n"
        else:
            report += "No feedback received.\n"
    
    report += f"""
{'='*70}
TONE SCORE (Client-Focus 1-10)
{'='*70}
"""
    tone_result = tone_result or {}
    ts = tone_result.get("tone_score")
    tn = tone_result.get("tone_notes", "")
    te = tone_result.get("error")
    if te:
        report += f"Error: {te}\n"
    elif ts is not None:
        report += f"Tone Score: {ts}/10\n"
        if tn:
            # Avoid duplicating "Score: N" line
            cleaned = re.sub(r"^(?:Score:\s*)?\d+(?:/10)?\s*\n?", "", tn.strip())
            if cleaned:
                report += f"{cleaned[:400]}\n"
    else:
        report += "Tone score: not run\n"
    
    report += f"""
{'='*70}
TASK 4: SSL / SECURITY (SSL Labs)
{'='*70}
"""
    ssl_grade = task4_result.get("ssl_grade")
    ssl_err = task4_result.get("ssl_error")
    if ssl_grade:
        report += f"SSL Grade: {ssl_grade}\n"
    elif ssl_err:
        report += f"SSL Labs: {ssl_err}\n"
    else:
        report += "SSL: not run\n"
    
    report += f"""
{'='*70}
AXE ACCESSIBILITY (WCAG)
{'='*70}
"""
    axe_result = axe_result or {}
    av = axe_result.get("a11y_violations_count", 0)
    at5 = axe_result.get("a11y_top_5", "")
    ae = axe_result.get("error")
    if ae:
        report += f"Error: {ae}\n"
    else:
        report += f"Violations: {av}\n"
        if at5:
            report += f"Top issues: {at5}\n"
    
    report += f"""
{'='*70}
END OF REPORT
{'='*70}
"""
    
    return report


def save_report(report: str, url: str, timestamp: datetime) -> str:
    """Save the report to audits subfolder with timestamp format: time_date_audit.txt"""
    # Create audits subfolder if it doesn't exist
    audits_dir = Path("audits")
    audits_dir.mkdir(exist_ok=True)
    
    # Format: HHMM_DDMMYY_audit.txt (e.g., 1635_260126_audit.txt)
    time_str = timestamp.strftime("%H%M")
    date_str = timestamp.strftime("%d%m%y")
    filename = f"{time_str}_{date_str}_audit.txt"
    filepath = audits_dir / filename
    
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(report)
    
    return str(filepath.absolute())


def log_to_csv(url: str, timestamp: datetime, task1_result: dict, task2_result: dict, 
                task3_result: dict, ai_provider: str, task4_result: dict = None,
                tone_result: dict = None, axe_result: dict = None) -> None:
    """Log audit results to CSV or XLSX file (prefers XLSX if available) with enhanced + free-check fields"""
    task4_result = task4_result or {}
    tone_result = tone_result or {}
    axe_result = axe_result or {}
    
    # Check if xlsx file exists, use it if available
    xlsx_file = Path("audit_log.xlsx")
    csv_file = Path("audit_log.csv")
    
    # Extract data
    perf_score = task1_result.get("performance_score", "N/A")
    perf_status = task1_result.get("status", "N/A")
    a11y = task1_result.get("accessibility_score")
    a11y_str = a11y if a11y is not None else "N/A"
    crux = task1_result.get("crux") or {}
    crux_lcp = crux.get("lcp_p75") if crux and not crux.get("error") else ""
    crux_fid = crux.get("fid_p75") if crux and not crux.get("error") else ""
    crux_cls = crux.get("cls_p75") if crux and not crux.get("error") else ""
    jsonld = "Yes" if task2_result.get("jsonld_exists") else "No"
    schema_types = ", ".join(task2_result.get("schema_types_found", [])) or "None"
    title = task2_result.get("title", "Not found")
    business_name = task2_result.get("business_name", "")
    email = task2_result.get("email", "")
    google_maps_link = task2_result.get("google_maps_link", "")
    city = task2_result.get("city", "")
    region = task2_result.get("region", "")
    ai_visibility_score = task2_result.get("ai_visibility_score", 0)
    ai_visibility_max = task2_result.get("ai_visibility_max", 5)
    ai_visibility_status = "EXCELLENT" if ai_visibility_score >= 4 else ("BASIC" if ai_visibility_score >= 2 else "POOR")
    missing_opportunities = "; ".join(task2_result.get("missing_opportunities", [])) or "None"
    sh = task2_result.get("security_headers")
    security_pass = "Yes" if isinstance(sh, dict) and sh.get("pass") else "No"
    security_missing = "; ".join(sh.get("missing", [])[:5]) if isinstance(sh, dict) else ""
    gtm_present = "Yes" if task2_result.get("gtm_present") else "No"
    gtm_id = task2_result.get("gtm_id", "")
    ga4_present = "Yes" if task2_result.get("ga4_present") else "No"
    ga4_id = task2_result.get("ga4_id", "")
    clarity_present = "Yes" if task2_result.get("clarity_present") else "No"
    schema_valid = "Yes" if task2_result.get("schema_valid") else "No"
    schema_warnings = "; ".join((task2_result.get("schema_warnings") or [])[:3])
    ssl_grade = task4_result.get("ssl_grade") or ""
    ssl_error = task4_result.get("ssl_error") or ""
    ai_feedback = "Yes" if task3_result.get("conversion_feedback") else "No"
    tone_score = tone_result.get("tone_score")
    tone_notes = (tone_result.get("tone_notes") or "")[:300]
    a11y_count = axe_result.get("a11y_violations_count", 0)
    a11y_top5 = (axe_result.get("a11y_top_5") or "")[:200]
    cms = task2_result.get("cms", "")
    analytics_tool = task2_result.get("analytics_tool", "")
    chat_widget = task2_result.get("chat_widget", "")
    framework = task2_result.get("framework", "")
    tech_stack = (task2_result.get("tech_stack_summary") or "")[:150]
    
    # Collect errors
    errors = []
    if task1_result.get("error"):
        errors.append(f"Task1: {task1_result['error'][:50]}")
    if task2_result.get("error"):
        errors.append(f"Task2: {task2_result['error'][:50]}")
    if task3_result.get("error"):
        errors.append(f"Task3: {task3_result['error'][:50]}")
    error_str = "; ".join(errors) if errors else "None"
    
    # Enhanced headers including tone, axe, wappalyzer
    headers = [
        "Timestamp", "URL", "Business_Name", "Performance_Score", "Performance_Status",
        "Accessibility_Score", "CrUX_LCP", "CrUX_FID", "CrUX_CLS",
        "JSONLD_Exists", "Schema_Types_Found", "AI_Visibility_Score", "AI_Visibility_Max",
        "AI_Visibility_Status", "Missing_Opportunities", "Email", "Google_Maps_Link",
        "City", "Region", "Title",
        "Security_Headers_Pass", "Security_Headers_Missing", "GTM_Present", "GTM_ID",
        "GA4_Present", "GA4_ID", "Clarity_Present", "Schema_Valid", "Schema_Warnings",
        "CMS", "Analytics_Tool", "Chat_Widget", "Framework", "Tech_Stack_Summary",
        "Tone_Score", "Tone_Notes", "A11y_Violations_Count", "A11y_Top_5_Issues",
        "SSL_Grade", "SSL_Error",
        "AI_Provider", "AI_Feedback_Received", "Errors"
    ]
    
    row_data = [
        timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        url,
        business_name,
        perf_score,
        perf_status,
        a11y_str,
        crux_lcp,
        crux_fid,
        crux_cls,
        jsonld,
        schema_types,
        ai_visibility_score,
        ai_visibility_max,
        ai_visibility_status,
        missing_opportunities,
        email,
        google_maps_link,
        city,
        region,
        title,
        security_pass,
        security_missing,
        gtm_present,
        gtm_id,
        ga4_present,
        ga4_id,
        clarity_present,
        schema_valid,
        schema_warnings,
        cms,
        analytics_tool,
        chat_widget,
        framework,
        tech_stack,
        tone_score if tone_score is not None else "",
        tone_notes,
        a11y_count,
        a11y_top5,
        ssl_grade,
        ssl_error,
        ai_provider,
        ai_feedback,
        error_str
    ]
    
    # Try to use XLSX if file exists and openpyxl is available
    if xlsx_file.exists() and OPENPYXL_AVAILABLE:
        try:
            wb = load_workbook(xlsx_file)
            ws = wb.active
            
            # Check if header row exists (first row should have "Timestamp")
            if ws.max_row == 0 or ws.cell(row=1, column=1).value != "Timestamp":
                ws.append(headers)
            else:
                # Migration: if we have new columns, update header row
                existing = [c.value for c in ws[1]]
                if len(existing) < len(headers):
                    for col, h in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=h)
            
            ws.append(row_data)
            wb.save(xlsx_file)
            return
        except Exception as e:
            print(f"  ⚠ Warning: Could not write to XLSX file: {e}")
            print(f"  → Falling back to CSV format")
    
    # Fall back to CSV
    file_exists = csv_file.exists()
    with open(csv_file, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        
        # Write header if file is new
        if not file_exists:
            writer.writerow(headers)
        
        # Write row
        writer.writerow(row_data)


def run_audit(
    url: str,
    *,
    skip_ssl: bool = True,
    no_cache: bool = True,
    quiet: bool = False,
) -> dict:
    """
    Run full audit and return structured results (for API / UI use).
    Uses env vars for API keys only; no interactive prompts.
    """
    import io
    from contextlib import redirect_stdout, redirect_stderr

    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    pagespeed_key = os.getenv("PAGESPEED_API_KEY")
    mistral_key = os.getenv("MISTRAL_API_KEY")
    xai_key = os.getenv("XAI_API_KEY")
    gemini_key = os.getenv("GEMINI_API_KEY")

    if not pagespeed_key or (not mistral_key and not xai_key and not gemini_key):
        return {
            "ok": False,
            "error": "Missing API keys. Set PAGESPEED_API_KEY and one of XAI_API_KEY, MISTRAL_API_KEY, GEMINI_API_KEY in .env",
            "url": url,
        }

    use_xai = bool(xai_key)
    use_mistral = bool(mistral_key) and not use_xai
    use_gemini = bool(gemini_key) and not use_xai and not use_mistral
    if use_mistral:
        ai_api_key = mistral_key
        ai_provider = "Mistral AI"
    elif use_xai:
        ai_api_key = xai_key
        ai_provider = "XAI/Grok"
    else:
        ai_api_key = gemini_key
        ai_provider = "Gemini"

    def _run():
        task1_result = task1_performance_check(url, pagespeed_key)
        task2_result = task2_technical_check(url)
        if use_mistral:
            task3_result = task3_ai_conversion_audit_mistral(url, ai_api_key)
        elif use_xai:
            task3_result = task3_ai_conversion_audit_xai(url, ai_api_key)
        else:
            task3_result = task3_ai_conversion_audit_gemini(url, ai_api_key)
        provider_key = "xai" if use_xai else ("mistral" if use_mistral else "gemini")
        tone_result = task_tone_score(url, ai_api_key, provider_key)
        axe_result = task_axe_accessibility(url)
        task4_result = {"ssl_grade": None, "ssl_error": "skipped"}
        if not skip_ssl:
            host = _get_hostname(url)
            if host:
                task4_result = task4_ssl_security(host, timeout_s=60)
        report = generate_report(url, task1_result, task2_result, task3_result, task4_result, tone_result, axe_result)
        return {
            "ok": True,
            "url": url,
            "task1": task1_result,
            "task2": task2_result,
            "task3": task3_result,
            "tone_result": tone_result,
            "axe_result": axe_result,
            "task4": task4_result,
            "report": report,
            "ai_provider": ai_provider,
        }

    if quiet:
        out = io.StringIO()
        err = io.StringIO()
        with redirect_stdout(out), redirect_stderr(err):
            return _run()
    return _run()


def main():
    """Main CLI entry point."""
    parser = argparse.ArgumentParser(
        description="SiteAuditBot - Comprehensive Sales Audit Report Generator"
    )
    parser.add_argument(
        "url",
        type=str,
        help="The URL to audit"
    )
    parser.add_argument(
        "--pagespeed-key",
        type=str,
        help="Google PageSpeed Insights API Key (or set PAGESPEED_API_KEY env var)"
    )
    parser.add_argument(
        "--mistral-key",
        type=str,
        help="Mistral AI API Key (or set MISTRAL_API_KEY env var)"
    )
    parser.add_argument(
        "--xai-key",
        type=str,
        help="XAI/Grok API Key (or set XAI_API_KEY env var)"
    )
    parser.add_argument(
        "--gemini-key",
        type=str,
        help="Google Gemini API Key (or set GEMINI_API_KEY env var)"
    )
    parser.add_argument(
        "--cache-ttl-hours",
        type=float,
        default=168.0,
        help="Cache TTL in hours (default: 168h = 7 days). Set to 0 to disable caching."
    )
    parser.add_argument(
        "--no-cache",
        action="store_true",
        help="Disable caching for this run"
    )
    parser.add_argument(
        "--skip-ssl",
        action="store_true",
        help="Skip SSL Labs check (saves ~60s)"
    )
    
    args = parser.parse_args()
    
    # Validate URL
    if not args.url.startswith(("http://", "https://")):
        args.url = "https://" + args.url
    
    print(f"\n🔍 SiteAuditBot - Starting audit for: {args.url}")
    print("="*70)
    
    # Get API keys
    pagespeed_key = args.pagespeed_key or os.getenv("PAGESPEED_API_KEY")
    mistral_key = args.mistral_key or os.getenv("MISTRAL_API_KEY")
    xai_key = args.xai_key or os.getenv("XAI_API_KEY")
    gemini_key = args.gemini_key or os.getenv("GEMINI_API_KEY")
    
    if not pagespeed_key or (not mistral_key and not xai_key and not gemini_key):
        pagespeed_key, mistral_key, xai_key, gemini_key = get_api_keys()
    
    # Determine which AI API to use (prefer XAI if specified, then Mistral, then Gemini)
    # For this run, prioritize XAI over Mistral
    use_xai = bool(xai_key)
    use_mistral = bool(mistral_key) and not use_xai
    use_gemini = bool(gemini_key) and not use_xai and not use_mistral
    
    if use_mistral:
        ai_api_key = mistral_key
        ai_provider = "Mistral AI"
    elif use_xai:
        ai_api_key = xai_key
        ai_provider = "XAI/Grok"
    elif use_gemini:
        ai_api_key = gemini_key
        ai_provider = "Gemini"
    else:
        print("✗ Error: No AI API key provided (Mistral, XAI, or Gemini required)")
        return 1
    
    print(f"✓ Using {ai_provider} for AI conversion audit")
    
    # Load cache if enabled
    cache_path = Path("audit_cache.json")
    cache_enabled = not args.no_cache and args.cache_ttl_hours > 0
    if cache_enabled:
        global audit_cache
        audit_cache = _load_audit_cache(str(cache_path))
        cache_key = args.url
        cached = _cache_get(audit_cache, cache_key, args.cache_ttl_hours * 3600.0)
        if cached:
            print(f"\n✓ Using cached results (from {datetime.fromtimestamp(cached.get('ts', 0)).strftime('%Y-%m-%d %H:%M:%S')})")
            task1_result = cached.get("task1", {})
            task2_result = cached.get("task2", {})
            task3_result = cached.get("task3", {})
        else:
            # Run all three tasks
            task1_result = task1_performance_check(args.url, pagespeed_key)
            task2_result = task2_technical_check(args.url)
            
            # Use the selected AI API
            if use_mistral:
                task3_result = task3_ai_conversion_audit_mistral(args.url, ai_api_key)
            elif use_xai:
                task3_result = task3_ai_conversion_audit_xai(args.url, ai_api_key)
            else:
                task3_result = task3_ai_conversion_audit_gemini(args.url, ai_api_key)
            
            # Save to cache
            with cache_lock:
                audit_cache[cache_key] = {
                    "v": AUDIT_CACHE_VERSION,
                    "ts": time.time(),
                    "task1": task1_result,
                    "task2": task2_result,
                    "task3": task3_result,
                }
                _save_audit_cache(str(cache_path), audit_cache)
    else:
        # Run all three tasks (no cache)
        task1_result = task1_performance_check(args.url, pagespeed_key)
        task2_result = task2_technical_check(args.url)
        
        # Use the selected AI API
        if use_mistral:
            task3_result = task3_ai_conversion_audit_mistral(args.url, ai_api_key)
        elif use_xai:
            task3_result = task3_ai_conversion_audit_xai(args.url, ai_api_key)
        else:
            task3_result = task3_ai_conversion_audit_gemini(args.url, ai_api_key)
    
    # Tone Score (AI 1-10 client-focus). Same provider as Task 3.
    provider_key = "xai" if use_xai else ("mistral" if use_mistral else "gemini")
    print("\n[Tone Score] Running AI Tone / Client-Focus (1-10)...")
    tone_result = task_tone_score(args.url, ai_api_key, provider_key)
    if tone_result.get("tone_score") is not None:
        print(f"  ✓ Tone Score: {tone_result['tone_score']}/10")
    elif tone_result.get("error"):
        print(f"  ⚠ Tone: {tone_result['error']}")
    
    # Axe Accessibility (WCAG). Always run.
    print("\n[Axe] Running axe-core accessibility...")
    axe_result = task_axe_accessibility(args.url)
    if axe_result.get("error"):
        print(f"  ⚠ Axe: {axe_result['error']}")
    else:
        print(f"  ✓ Violations: {axe_result.get('a11y_violations_count', 0)}")
    
    # Task 4: SSL Labs (free). Always run unless --skip-ssl.
    task4_result = {}
    if not getattr(args, "skip_ssl", False):
        host = _get_hostname(args.url)
        if host:
            task4_result = task4_ssl_security(host, timeout_s=60)
    else:
        task4_result = {"ssl_grade": None, "ssl_error": "skipped"}
    
    # Generate and display report
    print("\n" + "="*70)
    print("AUDIT COMPLETE - Generating Report...")
    print("="*70)
    
    timestamp = datetime.now()
    report = generate_report(args.url, task1_result, task2_result, task3_result, task4_result, tone_result, axe_result)
    print(report)
    
    # Save report to audits subfolder
    filepath = save_report(report, args.url, timestamp)
    print(f"\n✅ Report saved to: {filepath}")
    
    # Log to CSV/XLSX
    log_to_csv(args.url, timestamp, task1_result, task2_result, task3_result, ai_provider, task4_result, tone_result, axe_result)
    
    # Determine which file was used
    xlsx_file = Path("audit_log.xlsx")
    csv_file = Path("audit_log.csv")
    if xlsx_file.exists() and OPENPYXL_AVAILABLE:
        print(f"✅ Logged to: audit_log.xlsx\n")
    else:
        print(f"✅ Logged to: audit_log.csv\n")
    
    return 0


if __name__ == "__main__":
    exit(main())
